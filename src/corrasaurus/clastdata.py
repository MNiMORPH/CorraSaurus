"""Build per-site clast observations (number / area / mass fractions) from raw data.

Reads the raw per-clast field data (ClastCountCompilation.xlsx: one sheet per
site, one row per clast with a size and a field lithology code) and the
editable, field-based translation table (data/lithology_code_map.csv) that maps
each field code to one of the five source categories (or excludes it).

**Wolman sampling and the three fractions.**  The Toro counts are Wolman
blind-finger point counts: step, point at the bed without looking, pick the
clast under the finger.  The probability of picking a given clast is therefore
proportional to its exposed cross-section, ``~ D^2``, so the raw tally is an
*area-biased* sample, not a by-number one.  Nothing equals the raw tally; the
raw tally is an unbiased estimator of the bed's **area** fraction.

To recover any *bed* fraction we apply the Horvitz-Thompson correction, weighting
each sampled clast by ``1 / P(pick) ~ 1 / D^2`` and then by the moment of the
target quantity (``D^0`` number, ``D^2`` area, ``D^3`` mass/volume).  The net
per-clast weight is ``D^(p-2)``:

* **number fraction** -- weight ``1 / D^2``  (``Sum 1/D^2``)
* **area   fraction** -- weight ``1``        (``Sum 1`` = the raw tally)
* **mass   fraction** -- weight ``D``        (``Sum D``;  *not* ``Sum D^3``,
  which would double-count the D^2 sampling bias)

Sternberg's law is a mass-loss law, so the mass fraction is the physically
appropriate observable; the area fraction (= raw tally) is what the count
channel compares against directly, and the number fraction is kept for
diagnostics.

**Clast shape.**  The weights above take the measured "size" (the intermediate,
b-axis) as a clast diameter, which is exact only for equant clasts.  For a platy
clast lying flat (axes a >= b >= c) the projected area is ``a*b`` and the volume
``a*b*c``, so the Horvitz-Thompson **mass** weight is ``volume / projected_area =
c`` -- the *thickness*, not the b-axis -- and the **number** weight picks up
``b/a``.  Crucially the **area** fraction is shape-robust (it equals the raw
tally regardless of shape, because Wolman sampling *is* by projected area).  We
do not have all three axes, so shape enters as one constant per lithology via
``shape_factors``: the mass weight is scaled by ``c/b`` and the number weight by
``b/a`` (both 1 for an equant clast).  Toro schist is platy (~4:2:1 a:b:c),
so ``c/b = 0.5`` cuts its mass fraction ~2x relative to the equant assumption.

The code map is the source of truth: edit the CSV, not this module.
"""

from __future__ import annotations

import csv

import numpy as np
import pandas as pd

# Per-clast rows in each sheet start at row 9; col B = size (mm), col C = code.
_FIRST_DATA_ROW = 9
_SIZE_COL = 1   # 0-based within (size, code) we read cols B,C
_CODE_COL = 2


def load_code_map(csv_path: str) -> dict:
    """Read data/lithology_code_map.csv -> {field_code: lith_index}.

    Only rows with an integer ``lith_index`` are returned; excluded and
    UNASSIGNED codes are omitted (so callers treat them as "not a source").
    """
    code_map = {}
    with open(csv_path, newline="") as f:
        for row in csv.DictReader(f):
            idx = row.get("lith_index", "").strip()
            if idx.isdigit():
                code_map[row["code"].strip()] = int(idx)
    return code_map


def build_observations(xlsx_path: str, code_map: dict, categories, sites=None,
                       shape_factors=None) -> pd.DataFrame:
    """Per-site Horvitz-Thompson number / area / mass fractions.

    Each clast in the (area-biased, ~D^2) Wolman sample is weighted to recover
    the corresponding *bed* fraction: ``1/D^2`` for number, ``1`` for area
    (= the raw tally), ``D`` for mass.  See the module docstring.

    Parameters
    ----------
    xlsx_path : str
        ClastCountCompilation.xlsx (one sheet per site).
    code_map : dict
        field-code -> lith_index (from :func:`load_code_map`).
    categories : Categories
        ordered class set; defines column order and the lith_index->position
        and name->position maps.
    sites : iterable of str, optional
        Restrict to these sheet names; default = all sheets except "Template".
    shape_factors : dict, optional
        Per-lithology clast-shape correction, ``{lith_name: {"c_b": .., "b_a": ..}}``
        with ``c_b`` = shortest/intermediate (scales the mass weight) and ``b_a`` =
        intermediate/longest (scales the number weight).  Omitted lithologies are
        equant (factors 1).  The area fraction is shape-robust and never scaled.

    Returns
    -------
    DataFrame indexed by site with columns:
        n_clasts, n_unassigned,
        number_frac_<name>..., area_frac_<name>..., mass_frac_<name>...
        (in ``categories`` order)
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if sites is None:
        sites = [s for s in wb.sheetnames if s != "Template"]

    n_liths = len(categories)
    pbi = categories.position_by_index
    pbn = categories.position_by_name
    # per-lithology shape factors (default equant): mass *= c/b, number *= b/a.
    mass_shape = np.ones(n_liths)
    number_shape = np.ones(n_liths)
    for name, sf in (shape_factors or {}).items():
        if name not in pbn:
            raise ValueError(f"unknown lithology in shape_factors: {name!r}")
        p = pbn[name]
        mass_shape[p] = float(sf.get("c_b", 1.0))
        number_shape[p] = float(sf.get("b_a", 1.0))

    rows = {}
    for site in sites:
        if site not in wb.sheetnames:
            continue
        ws = wb[site]
        n_count = np.zeros(n_liths)   # raw clast tally (per lithology)
        s_number = np.zeros(n_liths)  # Sum 1/D^2  -> bed number fraction
        s_area = np.zeros(n_liths)    # Sum 1      -> bed area fraction (= raw tally)
        s_mass = np.zeros(n_liths)    # Sum D      -> bed mass fraction
        n_unassigned = 0
        for r in ws.iter_rows(min_row=_FIRST_DATA_ROW, max_col=3, values_only=True):
            size, code = r[_SIZE_COL], r[_CODE_COL]
            if size is None or code is None:
                continue
            code = str(code).strip()
            if code in ("", "lithology"):
                continue
            idx = code_map.get(code)
            if idx is None:
                n_unassigned += 1
                continue
            D = float(size)
            if D <= 0:   # cannot area-sample (P ~ D^2) a non-positive clast size
                continue
            pos = pbi[idx]
            n_count[pos] += 1
            s_number[pos] += 1.0 / D ** 2   # Horvitz-Thompson: number ~ D^(0-2)
            s_area[pos] += 1.0              #                    area   ~ D^(2-2)
            s_mass[pos] += D                #                    mass   ~ D^(3-2)
        tot_c = n_count.sum()
        if tot_c == 0:
            continue
        s_number = s_number * number_shape   # b/a (area fraction stays shape-robust)
        s_mass = s_mass * mass_shape         # c/b (HT mass weight is the thickness)
        rec = {"n_clasts": int(tot_c), "n_unassigned": n_unassigned}
        tot_number, tot_area, tot_mass = s_number.sum(), s_area.sum(), s_mass.sum()
        for i, name in enumerate(categories.names):
            rec[f"number_frac_{name}"] = s_number[i] / tot_number
            rec[f"area_frac_{name}"] = s_area[i] / tot_area
            rec[f"mass_frac_{name}"] = s_mass[i] / tot_mass
        rows[site] = rec

    df = pd.DataFrame.from_dict(rows, orient="index")
    df.index.name = "site"
    return df


def load_clast_records(xlsx_path: str, code_map: dict, categories, sites) -> dict:
    """Per-clast records for resampling (the nonparametric bootstrap).

    Returns ``{site: (lith_pos, size_mm)}`` where ``lith_pos`` is the canonical
    lithology position (0..n_liths-1) and ``size_mm`` the clast size, for every
    mapped clast.  Keeping the raw clasts (not just aggregated fractions) lets a
    bootstrap resample the ~100 measured clasts per site -- propagating both the
    finite-count sampling and (for mass) the dominance of the largest clasts.
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    out = {}
    pbi = categories.position_by_index
    for site in sites:
        if site not in wb.sheetnames:
            continue
        ws = wb[site]
        liths, sizes = [], []
        for r in ws.iter_rows(min_row=_FIRST_DATA_ROW, max_col=3, values_only=True):
            size, code = r[_SIZE_COL], r[_CODE_COL]
            if size is None or code is None:
                continue
            idx = code_map.get(str(code).strip())
            if idx is None:
                continue
            liths.append(pbi[idx])
            sizes.append(float(size))
        out[site] = (np.array(liths, dtype=np.intp), np.array(sizes, dtype=float))
    return out


def size_moments(records: dict, categories, sites) -> tuple:
    """Per-(site, lithology) mean ln(size) and clast count, for the size channel.

    Downstream size-fining constrains attrition length directly (Sternberg is a
    mass/size law), independently of production -- the observable that lets the
    joint inversion separate the production-attrition degeneracy.

    Returns
    -------
    mean_lnD : array (n_sites, n_liths)
        Mean of ln(clast size [mm]) over each lithology's clasts at each site;
        NaN where a lithology has no clasts at a site.
    count : array (n_sites, n_liths)
        Number of clasts of each lithology at each site (the residual weight).
    """
    nl = len(categories)
    sites = list(sites)
    mean_lnD = np.full((len(sites), nl), np.nan)
    count = np.zeros((len(sites), nl))
    for i, s in enumerate(sites):
        lith, size = records[s]
        for k in range(nl):
            m = (lith == k) & (size > 0)
            n = int(m.sum())
            if n > 0:
                mean_lnD[i, k] = np.log(size[m]).mean()
                count[i, k] = n
    return mean_lnD, count


def fractions_matrix(df: pd.DataFrame, categories, kind: str = "mass") -> np.ndarray:
    """Extract the (n_sites, n_liths) fraction matrix in canonical order.

    kind : "mass", "area", or "number".  Legacy "count" is accepted as an alias
    for "area" -- a Wolman tally estimates the bed *area* fraction, not a
    by-number one.
    """
    kind = {"count": "area"}.get(kind, kind)
    if kind not in ("mass", "area", "number"):
        raise ValueError("kind must be 'mass', 'area', or 'number' "
                         "('count' is accepted as an alias for 'area')")
    cols = [f"{kind}_frac_{name}" for name in categories.names]
    return df[cols].to_numpy(dtype=float)
